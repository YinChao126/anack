# -*- coding: utf-8 -*-
"""
Created on Thu Jul 12 20:17:19 2018

@author: John
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re
import time

# read txt data


def read_txt(inputfiles):
    p1 = r"(.*)[0-9](.*?).*"
    temp_list = []
    with open(inputfiles,'r',encoding='utf-8') as f:
        for line in f:
            mathObj = re.match(p1,line)
            if mathObj:
                line = line.strip()
                temp_list.append(line)
    return temp_list



def parse_line(aline):
    aline = aline.replace(":"," ")
    aline = aline.replace("："," ")
    aline = aline.replace(","," ")
    aline = aline.split()
    return aline  
    

# create excel files
def generate_excel(temp_list,inputfiles):
    rows = len(temp_list)
    
    wb = Workbook()
    
    dest_filename = inputfiles[:-4] + '.xlsx'
    ws1 = wb.active
    
    ws1.title = "Analysis report V1"
    
    for row in range(rows):
        aline = temp_list[row]
        aline = parse_line(aline)
        for col in range(len(aline)):
            ws1.cell(column=col + 1, row=row + 1, value="{0}".format(aline[col]))
    
    wb.save(filename = dest_filename)

if __name__ == "__main__":
    inputfiles = r"D:\600522_20180714.txt"
    file_list = read_txt(inputfiles)
    generate_excel(file_list,inputfiles)